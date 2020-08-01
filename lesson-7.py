'''
python自动化工作：
1、准备好自动化测试用例  ==== done   test_case_api.xlsx
2、使用python去读取测试用例  ==== done   read_data  ---6
3、发送请求，得到响应结果 ====done     api_func  ---5
4、结果的判断，执行结果  vs  预期结果====断言
5、得到一个最终结果，回写到测试用例 ==== done    write_result ---6
'''
'''返回的结果（字典格式保存的）
'case_id': 3, 
'url': 'http://api.lemonban.com/futureloan/member/register', 
'data': '{"pwd":"12345678","type":1}', 
'expected': '{"code":1,"msg":"手机号为空"}'}
'''
import requests
import openpyxl
# 做完整的测试
# 1、读取测试用例
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  #改名为
    sheet = wb[sheetname]
    max_row = sheet.max_row #取出sheet里面最大的行数
    # print(max_row)
    case_list = []
    for i in range(2,max_row+1,1):
        dict1 = dict(
        case_id=sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value ,
        expected = sheet.cell(row=i,column=7).value)
        case_list.append(dict1)#把dict1里面一条一条的测试用例装到列表里面，这个列表里面存放了所有测试用例
    # print(case_list)
    return case_list

# 2、发送请求  要字典格式
def api_func(url,data):
  header_login = {'X-Lemonban-Media-Type': 'lemonban.v2',
                  'Content-Type': 'application/json'}  # 请求头
  res1 = requests.post(url=url, json=data, headers=header_login)
  # print(res1.json())  # 登录
  response = res1.json()
  return response

# 3、写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

# 4、读取数据
# cases= read_data('test_case_api.xlsx','register')   #读取Excel中的测试用例
# # print(cases)
#
# # 5、取值，（因为测试不是全部用例一起用）
# for case in cases:    #循环取出测试用例
#     # print(case)  #先打印这一步，在通过打印的结果取下面的值  6 (第六步）
#     case_id = case.get('case_id') #字典取值              6
#     url = case['url']  #取url是通过key的方法取值           6
#     data = case.get('data')  #通过Excel取出来的值是str类型的   6
#     # print(type(data))      #查看data是什么类型                #8
#     data = eval(data)  #eval()的作用：运行被字符串包括的python表达式
#     expected = case.get('expected')                     #9
#     expected = eval(expected)                          #11
#     expected_msg = expected.get('msg') #取出预期结果里的msg信息   11
#     # print(case_id,url,data,expected) #打印结果         6
#     # api_func(url, data) #调用了发送请求的函数  这一步不要打开   #7
#     real_result = api_func(url=url, data=data)  #调用了发送请求的函数并传入参数   7
#     # print(real_result)  #打印出来是密码为空                              7
#     # print(real_result)     #   打印出来账号已存在           10、
#     real_msg = real_result.get('msg')   #取出实际执行结果里面的msg信息    12
#     print('预期结果为：{}'.format(expected_msg))
#     print('实际结果为：{}'.format(real_msg))
#     if real_msg ==expected_msg:
#         print('第{}条用例通过!'.format(case_id))
#         final_res = 'pass'
#     else:
#         print('第{}条用例未通过!'.format(case_id))
#         final_res = 'fail'           #14
#     print('*' * 30)    #分层符号                        13
#     write_result('test_case_api.xlsx','register',case_id+1,8,final_res)

# 最后封装成函数
def execute_func(filename,sheetname):
    cases= read_data(filename,sheetname)   #读取Excel中的测试用例
    # print(cases)

    # 5、取值，（因为测试不是全部用例一起用）
    for case in cases:    #循环取出测试用例
        # print(case)  #先打印这一步，在通过打印的结果取下面的值  6 (第六步）
        case_id = case.get('case_id') #字典取值              6
        url = case['url']  #取url是通过key的方法取值           6
        data = case.get('data')  #通过Excel取出来的值是str类型的   6
        # print(type(data))      #查看data是什么类型                #8
        data = eval(data)  #eval()的作用：运行被字符串包括的python表达式
        expected = case.get('expected')                     #9
        expected = eval(expected)                          #11
        expected_msg = expected.get('msg') #取出预期结果里的msg信息   11
        # print(case_id,url,data,expected) #打印结果         6
        # api_func(url, data) #调用了发送请求的函数  这一步不要打开   #7
        real_result = api_func(url=url, data=data)  #调用了发送请求的函数并传入参数   7
        # print(real_result)  #打印出来是密码为空                              7
        # print(real_result)     #   打印出来账号已存在           10、
        real_msg = real_result.get('msg')   #取出实际执行结果里面的msg信息    12
        print('预期结果为：{}'.format(expected_msg))
        print('实际结果为：{}'.format(real_msg))
        if real_msg ==expected_msg:
            print('第{}条用例通过!'.format(case_id))
            final_res = 'pass'
        else:
            print('第{}条用例未通过!'.format(case_id))
            final_res = 'fail'           #14
        print('*' * 30)    #分层符号                        13
        write_result(filename,sheetname,case_id+1,8,final_res)
execute_func('test_case_api.xlsx','login')